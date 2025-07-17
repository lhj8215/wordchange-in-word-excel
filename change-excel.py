import os
import sys
from openpyxl import load_workbook

def replace_keyword_in_excel(old_keyword, new_keyword):
    folder_path = os.getcwd()
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            workbook = load_workbook(file_path)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and old_keyword in cell.value:
                            cell.value = cell.value.replace(old_keyword, new_keyword)
            workbook.save(file_path)
            print(f"Processed: {filename}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python change-excel.py <old_keyword> <new_keyword>")
        sys.exit(1)

    old_keyword = sys.argv[1]
    new_keyword = sys.argv[2]

    replace_keyword_in_excel(old_keyword, new_keyword)